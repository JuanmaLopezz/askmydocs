from pathlib import Path
from langchain_core.documents import Document
from langchain_community.document_loaders import PyPDFLoader, Docx2txtLoader
import openpyxl


def load_document(file_path: Path) -> list[Document]:
    ext = file_path.suffix.lower()
    if ext == ".pdf":
        return PyPDFLoader(str(file_path)).load()
    elif ext == ".docx":
        return Docx2txtLoader(str(file_path)).load()
    elif ext == ".xlsx":
        return _load_xlsx(file_path)
    elif ext == ".txt":
        return _load_txt(file_path)
    else:
        raise ValueError(f"Extension not supported: {ext}")


def _load_xlsx(file_path: Path) -> list[Document]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    docs = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [
            "\t".join(str(cell) if cell is not None else "" for cell in row)
            for row in ws.iter_rows(values_only=True)
        ]
        content = "\n".join(r for r in rows if r.strip())
        if content:
            docs.append(Document(
                page_content=content,
                metadata={"source": str(file_path), "sheet": sheet_name},
            ))
    return docs


def _load_txt(file_path: Path) -> list[Document]:
    text = file_path.read_text(encoding="utf-8", errors="ignore")
    return [Document(page_content=text, metadata={"source": str(file_path)})]
